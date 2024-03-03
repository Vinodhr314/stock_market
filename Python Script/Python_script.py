import openpyxl
import os

# Replace 'data.xlsx' with the actual filename of your Excel sheet
filename = r'C:\Users\VINOD H R\Desktop\stock_market\Historical data\SWSOLAR.xlsx'
# List of filenames:
filenames = [
    r'C:\Users\VINOD H R\Desktop\stock_market\Historical data\ZOMATO.xlsx',
    r'C:\Users\VINOD H R\Desktop\stock_market\Historical data\IND_HOTEL.xlsx',
    r'C:\Users\VINOD H R\Desktop\stock_market\Historical data\SW_SOLAR.xlsx',
    r'C:\Users\VINOD H R\Desktop\stock_market\Historical data\M&M.xlsx'
]
for filename in filenames:
    try:
        wb = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        print(f"Error: File '{filename}' not found.")
        exit()

    # Get the active worksheet
    sheet = wb.active

    # Clear the content from column O to Z (inclusive)
    for row in sheet.iter_rows():
        for cell in row[14:]:  # Start from column O (index 14)
            cell.value = None

    # Insert new columns
    columns = [
        ('CLOSE-OPEN', 15),
        ('HIGH-OPEN', 16),
        ('LOW-OPEN', 17),
        ('TCLOSE-PCLOSE', 18),
        ('HIGH-LOW', 19)
    ]
    for name, index in columns:
        sheet.insert_cols(index)
        sheet.cell(row=1, column=index).value = name

    # Assuming your data starts in row 2 (modify if needed)
    for row in range(2, sheet.max_row + 1):
        # Iterate through each calculation and handle missing values consistently
        for col1, col2, target_col in [
            (8, 3, 15),  # Close-Open
            (4, 3, 16),  # High-Open
            (5, 3, 17),  # Low-Open
            (8, 6, 18),  # Close-Close
            (4, 5, 19)   # High-Low
        ]:
            value1 = sheet.cell(row=row, column=col1).value
            value2 = sheet.cell(row=row, column=col2).value
            if value1 is not None and value2 is not None:
                sheet.cell(row=row, column=target_col).value = value1 - value2
            else:
                sheet.cell(row=row, column=target_col).value = "NA"

    # Save the changes to the Excel file
    wb.save(filename)

    for row in range(2, sheet.max_row + 1):
        # Color negative values in column O with red
        cell = sheet.cell(row=row, column=15)  # Assuming column O is index 15
        if cell.value < 0:
            cell.font = openpyxl.styles.Font(color="FF0000")  # Red
            cell.fill = openpyxl.styles.PatternFill(bgColor="FF0000")  # Light red background

        # Color negative values in column R with blue
        cell = sheet.cell(row=row, column=18)  # Assuming column R is index 18
        if cell.value < 0:
            cell.font = openpyxl.styles.Font(color="0000FF")  # Blue
            cell.fill = openpyxl.styles.PatternFill(bgColor="D9D9F3")  # Light blue background

    wb.save(filename)  # Save the file again to apply colors

    print("Calculation completed and saved and opening now: ", filename)

    # Open the modified file using the appropriate method based on your operating system:
    if os.name == 'nt':  # Windows
        os.startfile(filename)
    elif os.name == 'posix':  # Linux or macOS
        subprocess.call(['open', filename])  # Replace 'open' with the appropriate command if needed
    else:
        print("Unsupported operating system. File not opened automatically.")